# Python-Software for counting punishment in Farms (Normalization)

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import re
import sys
from PIL import Image, ImageTk, ImageDraw, ImageChops


#--------------------------------------------------- SPLASH SCREEN ----------------------------------------------------#
def show_splash_and_start(main_window_callback):
    splash = tk.Tk()
    splash.overrideredirect(True)  # Χωρίς μπάρα παραθύρου (Χ, ελαχιστοποίηση)

    splash_width = 460
    splash_height = 360

    # Κεντράρισμα στην οθόνη
    screen_width = splash.winfo_screenwidth()
    screen_height = splash.winfo_screenheight()
    x = (screen_width / 2) - (splash_width / 2)
    y = (screen_height / 2) - (splash_height / 2)
    splash.geometry(f'{splash_width}x{splash_height}+{int(x)}+{int(y)}')

    # Κύριο background frame
    main_frame = tk.Frame(splash, bg="white")
    main_frame.pack(expand=True, fill="both")

    # Σταθερό frame για την εικόνα
    img_container = tk.Frame(main_frame, width=150, height=130, bg="white")
    img_container.pack(pady=(25, 10))
    img_container.pack_propagate(False)

    img_label = tk.Label(img_container, bg="white")
    img_label.place(relx=0.5, rely=0.5, anchor="center")

    try:
        ico_path = resource_path(LOGO_FILENAME)
        original_img = Image.open(ico_path).convert("RGBA")

        # Διατήρηση σωστής αναλογίας (Aspect Ratio)
        orig_w, orig_h = original_img.size
        max_size = 110
        ratio = min(max_size / orig_w, max_size / orig_h)
        new_w = int(orig_w * ratio)
        new_h = int(orig_h * ratio)

        resized_img = original_img.resize((new_w, new_h), Image.Resampling.LANCZOS)

        # ΝΕΟ ΕΦΕ: Fade με διασταυρούμενες γραμμές
        def fade_in_lines(progress=0.0):
            if progress <= 1.0:
                # Δημιουργία μιας μαύρης μάσκας
                mask = Image.new("L", resized_img.size, 0)
                draw = ImageDraw.Draw(mask)

                step = 10  # Μέγεθος κελιού του "πλέγματος". Μπορείς να το αλλάξεις (π.χ. 8 ή 12)
                thickness = int(step * progress)

                if thickness > 0:
                    # Σχεδιασμός οριζόντιων γραμμών
                    for y in range(0, new_h, step):
                        draw.rectangle([0, y, new_w, y + thickness], fill=int(255 * progress))

                    # Σχεδιασμός κάθετων γραμμών (για τη διασταύρωση)
                    for x in range(0, new_w, step):
                        draw.rectangle([x, 0, x + thickness, new_h], fill=int(255 * progress))

                faded_img = resized_img.copy()

                # Παίρνουμε το αρχικό Alpha (τη διαφάνεια) της εικόνας (αν έχει στρογγυλές άκρες κλπ)
                orig_alpha = faded_img.getchannel('A')

                # Πολλαπλασιάζουμε το αρχικό Alpha με τη μάσκα των γραμμών μας
                final_alpha = ImageChops.multiply(orig_alpha, mask)
                faded_img.putalpha(final_alpha)

                bg = Image.new("RGBA", faded_img.size, (255, 255, 255, 255))
                blended = Image.alpha_composite(bg, faded_img).convert("RGB")

                splash.photo = ImageTk.PhotoImage(blended)
                img_label.config(image=splash.photo)

                # Αυξάνουμε το progress (ταχύτητα εφέ)
                splash.after(40, lambda: fade_in_lines(progress + 0.05))

        splash.after(100, fade_in_lines)

    except Exception as e:
        print("Σφάλμα φόρτωσης εικόνας:", e)

    # Σταθερό frame για τα κείμενα
    text_container = tk.Frame(main_frame, bg="white")
    text_container.pack(fill="x", padx=20)

    # 1ο Κείμενο: Τίτλος συστήματος
    title_label = tk.Label(text_container, text="",
                           font=("Georgia", 16, "bold"), fg="#0d233a", bg="white",
                           anchor="center", justify="center")
    title_label.pack(fill="x", pady=(0, 4))

    # 2ο Κείμενο: Υπουργείο
    subtitle_label = tk.Label(text_container, text="",
                              font=("Segoe UI", 11, "bold"), fg="#1f4e78", bg="white",
                              anchor="center", justify="center")
    subtitle_label.pack(fill="x")

    full_title = "Σύστημα Βιοπροφύλαξης"
    full_subtitle = "ΥΠΟΥΡΓΕΙΟ ΑΓΡΟΤΙΚΗΣ ΑΝΑΠΤΥΞΗΣ & ΤΡΟΦΙΜΩΝ"

    def animate_title(index=0):
        if index <= len(full_title):
            title_label.config(text=full_title[:index])
            splash.after(45, lambda: animate_title(index + 1))
        else:
            animate_subtitle()

    def animate_subtitle(index=0):
        if index <= len(full_subtitle):
            subtitle_label.config(text=full_subtitle[:index])
            splash.after(25, lambda: animate_subtitle(index + 1))

    splash.after(400, animate_title)

    # Συνάρτηση αναμονής και κλεισίματος splash
    # ΣΗΜΑΝΤΙΚΟ: τρέχει μέσα στο ίδιο mainloop μέσω splash.after(), ΟΧΙ σε ξεχωριστό thread.
    # Το Tkinter δεν είναι thread-safe· destroy()/after() που καλούνταν από άλλο thread
    # ήταν η αιτία των σφαλμάτων "main thread is not in main loop" και
    # "application has been destroyed" (ThemeChanged), καθώς και της ασταθούς
    # συμπεριφοράς στα icons/animations.

    def close_splash_and_launch():
        curr_width = 460
        curr_height = 360

        # Διαβάζουμε την τρέχουσα θέση του παραθύρου στην οθόνη
        current_x = splash.winfo_x()
        current_y = splash.winfo_y()

        # Χωρίζουμε την απόσταση σε περίπου 20 "καρέ/βήματα" για ομαλή κίνηση
        step_x = current_x / 20.0
        step_y = current_y / 20.0

        def animate_to_top_left(alpha, w, h, x, y):
            if alpha > 0:
                alpha -= 0.05  # Σταδιακό σβήσιμο
                w -= 23  # Σμίκρυνση πλάτους (460 / 20)
                h -= 18  # Σμίκρυνση ύψους (360 / 20)

                x -= step_x  # Μετακίνηση προς τα αριστερά
                y -= step_y  # Μετακίνηση προς τα πάνω

                # Αποτρέπουμε τις αρνητικές τιμές που μπορεί να κρασάρουν το Tkinter
                if w < 1: w = 1
                if h < 1: h = 1
                if x < 0: x = 0
                if y < 0: y = 0

                # Εφαρμογή των αλλαγών
                splash.attributes('-alpha', alpha)
                splash.geometry(f'{int(w)}x{int(h)}+{int(x)}+{int(y)}')

                # Κλήση του επόμενου καρέ
                splash.after(20, lambda: animate_to_top_left(alpha, w, h, x, y))
            else:
                splash.destroy()
                if main_window_callback:
                    main_window_callback()

        # Έναρξη του animation
        animate_to_top_left(1.0, curr_width, curr_height, current_x, current_y)

    splash.after(3000, close_splash_and_launch)
    splash.mainloop()


#------------------------------------------ MAIN PROGRAM --------------------------------------------------------------#
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OPENPYXL_AVAILABLE = True
    OPENPYXL_IMPORT_ERROR = None
except ImportError as _e:
    OPENPYXL_AVAILABLE = False
    OPENPYXL_IMPORT_ERROR = str(_e)

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.graphics.shapes import Drawing, Rect, String, Line

    REPORTLAB_AVAILABLE = True
    REPORTLAB_IMPORT_ERROR = None
except ImportError as _e:
    REPORTLAB_AVAILABLE = False
    REPORTLAB_IMPORT_ERROR = str(_e)

try:
    from fontTools.ttLib import TTFont as _FT_TTFont

    FONTTOOLS_AVAILABLE = True
except ImportError:
    FONTTOOLS_AVAILABLE = False


#------------------------ ΒΟΗΘΗΤΙΚΕΣ ΣΥΝΑΡΤΗΣΕΙΣ (ΠΟΡΟΙ, ΓΡΑΜΜΑΤΟΣΕΙΡΕΣ PDF, ΧΡΩΜΑΤΙΚΟ LEGEND) ------------------------#
def resource_path(relative_path):
    """Επιστρέφει το σωστό absolute path για ένα bundled αρχείο."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)


ICON_FILENAME = "ΥΠΑΑΤ_ICON_Exe.ico"
LOGO_FILENAME = "Screenshot 2026-06-25 130650.png"

PDF_FONT_REGULAR_FILENAME = "DejaVuSans.ttf"
PDF_FONT_BOLD_FILENAME = "DejaVuSans-Bold.ttf"

PDF_FONT_NAME = "Helvetica"
PDF_FONT_NAME_BOLD = "Helvetica-Bold"
_PDF_FONTS_REGISTERED = False

PDF_SYMBOLS_SUPPORTED = True
PDF_FONT_SOURCE_RELIABLE = True
PDF_FONT_SOURCE_PATH = None


def _register_pdf_fonts():
    global PDF_FONT_NAME, PDF_FONT_NAME_BOLD, _PDF_FONTS_REGISTERED, PDF_SYMBOLS_SUPPORTED
    global PDF_FONT_SOURCE_RELIABLE, PDF_FONT_SOURCE_PATH
    if _PDF_FONTS_REGISTERED:
        return
    _PDF_FONTS_REGISTERED = True

    reliable_regular = [resource_path(PDF_FONT_REGULAR_FILENAME)]
    reliable_bold = [resource_path(PDF_FONT_BOLD_FILENAME)]

    try:
        import matplotlib
        mpl_fonts = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
        reliable_regular.append(os.path.join(mpl_fonts, "DejaVuSans.ttf"))
        reliable_bold.append(os.path.join(mpl_fonts, "DejaVuSans-Bold.ttf"))
    except Exception:
        pass

    reliable_regular.append("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    reliable_bold.append("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")

    windir = os.environ.get("WINDIR", r"C:\Windows")
    fallback_regular = [
        os.path.join(windir, "Fonts", "arial.ttf"),
        os.path.join(windir, "Fonts", "calibri.ttf"),
        "/Library/Fonts/Arial.ttf",
    ]
    fallback_bold = [
        os.path.join(windir, "Fonts", "arialbd.ttf"),
        os.path.join(windir, "Fonts", "calibrib.ttf"),
        "/Library/Fonts/Arial Bold.ttf",
    ]

    def _first_existing_verified(paths):
        return next((p for p in paths if p and os.path.isfile(p) and _font_covers_required_glyphs(p)), None)

    def _first_existing(paths):
        return next((p for p in paths if p and os.path.isfile(p)), None)

    reg_path = _first_existing_verified(reliable_regular)
    source_reliable = True

    if not reg_path:
        reg_path = _first_existing(reliable_regular)
        source_reliable = reg_path is not None

    if not reg_path:
        reg_path = _first_existing_verified(fallback_regular) or _first_existing(fallback_regular)
        source_reliable = False

    if not reg_path:
        PDF_FONT_SOURCE_RELIABLE = False
        return

    PDF_FONT_SOURCE_PATH = reg_path
    PDF_FONT_SOURCE_RELIABLE = source_reliable

    reg_dir = os.path.dirname(reg_path)
    all_bold_candidates = reliable_bold + fallback_bold
    bold_path = next(
        (p for p in all_bold_candidates if p and os.path.isfile(p) and os.path.dirname(p) == reg_dir), None
    )
    if not bold_path:
        bold_path = _first_existing(all_bold_candidates)

    try:
        pdfmetrics.registerFont(TTFont("AppUnicode", reg_path))
        PDF_FONT_NAME = "AppUnicode"
        if bold_path:
            pdfmetrics.registerFont(TTFont("AppUnicode-Bold", bold_path))
            PDF_FONT_NAME_BOLD = "AppUnicode-Bold"
        else:
            PDF_FONT_NAME_BOLD = "AppUnicode"
        PDF_SYMBOLS_SUPPORTED = _font_covers_required_glyphs(reg_path)
    except Exception:
        PDF_FONT_NAME = "Helvetica"
        PDF_FONT_NAME_BOLD = "Helvetica-Bold"
        PDF_SYMBOLS_SUPPORTED = False
        PDF_FONT_SOURCE_RELIABLE = False


REQUIRED_PDF_GLYPHS = "ΑΩαω"


def _font_covers_required_glyphs(font_path):
    if not FONTTOOLS_AVAILABLE:
        return True
    try:
        ft = _FT_TTFont(font_path, lazy=True)
        cmap = ft.getBestCmap() or {}
        ft.close()
        return all(ord(ch) in cmap for ch in REQUIRED_PDF_GLYPHS)
    except Exception:
        return True


def _lerp_hex(c1_hex, c2_hex, t):
    c1_hex, c2_hex = c1_hex.lstrip("#"), c2_hex.lstrip("#")
    r1, g1, b1 = int(c1_hex[0:2], 16), int(c1_hex[2:4], 16), int(c1_hex[4:6], 16)
    r2, g2, b2 = int(c2_hex[0:2], 16), int(c2_hex[2:4], 16), int(c2_hex[4:6], 16)
    t = max(0.0, min(1.0, t))
    return rl_colors.Color(
        (r1 + (r2 - r1) * t) / 255.0,
        (g1 + (g2 - g1) * t) / 255.0,
        (b1 + (b2 - b1) * t) / 255.0,
    )


def _build_legend_drawing(width_pts):
    bar_h = 7
    bottom_h = 8
    top_h = 8
    bar_y = bottom_h

    d = Drawing(width_pts, bar_h + bottom_h + top_h)

    n_steps = 100
    seg_w = width_pts / n_steps
    for i in range(n_steps):
        pct = (i + 0.5) / n_steps * 100
        if pct <= 30:
            color = _lerp_hex(GOOD_GREEN_BG, WARN_YELLOW_BG, pct / 30.0)
        elif pct <= 50:
            color = _lerp_hex(WARN_YELLOW_BG, ORANGE_BG, (pct - 30) / 20.0)
        elif pct <= 90:
            color = _lerp_hex(ORANGE_BG, BAD_RED_BG, (pct - 50) / 40.0)
        else:
            color = rl_colors.HexColor(BAD_RED_BG)
        d.add(Rect(i * seg_w, bar_y, seg_w + 0.6, bar_h, fillColor=color, strokeColor=None))

    d.add(Rect(0, bar_y, width_pts, bar_h, fillColor=None,
               strokeColor=rl_colors.HexColor("#999999"), strokeWidth=0.4))

    for pct_mark in (0, 30, 50, 90, 100):
        x = width_pts * (pct_mark / 100.0)
        x_clamped = min(max(x, 1), width_pts - 1)
        d.add(Line(x_clamped, bar_y, x_clamped, bar_y + bar_h,
                   strokeColor=rl_colors.HexColor("#666666"), strokeWidth=0.4))
        d.add(String(x_clamped, 1, f"{pct_mark}%", fontName=PDF_FONT_NAME, fontSize=5,
                     fillColor=rl_colors.HexColor("#555555"), textAnchor="middle"))

    category_labels = [
        (15, "OK", GOOD_GREEN_TXT),
        (40, "ΗΠΙΑ ΜΗ ΣΥΜΜ.", WARN_YELLOW_TXT),
        (70, "ΣΟΒΑΡΗ ΜΗ ΣΥΜΜ.", ORANGE_TXT),
        (95, "ΜΗ ΕΦΑΡΜΟΣΙΜΟ", BAD_RED_TXT),
    ]
    for pct_center, text, fg in category_labels:
        x = width_pts * (pct_center / 100.0)
        d.add(String(x, bar_y + bar_h + 1.5, text, fontName=PDF_FONT_NAME_BOLD, fontSize=5,
                     fillColor=rl_colors.HexColor(fg), textAnchor="middle"))

    return d


#------------------------------------------- ΔΕΔΟΜΕΝΑ CHECKLIST (SECTIONS) --------------------------------------------#
SECTIONS = [
    {
        "title": "1. Απαιτήσεις για τις υποδομές των εγκαταστάσεων",
        "items": [
            "Πλήρης περίφραξη.",
            "Θύρες με δυνατότητα ασφαλούς κλεισίματος.",
            "Καθορισμός καθαρών και μη καθαρών περιοχών.",
            "Σε λειτουργία απολυμαντική τάφρος στην είσοδο ή ισοδύναμο μέσο για την απολύμανση οχημάτων / επισκεπτών.",
            "Εξοπλισμός απολύμανσης υποδημάτων στην είσοδο των κτιρίων.",
            "Σύστημα ορθής διαχείριση κόπρους και αποβλήτων.",
            "Υλικά κατασκευής υποδομών που επιτρέπουν τον αποτελεσματικό καθαρισμό και την απολύμανση (π.χ. λείες επιφάνειες).",
            "Ύπαρξη κελιών καραντίνας.",
            "Ύπαρξη κελιών νοσηλείας.",
            "Ειδικός χώρος στάθμευσης οχημάτων εξωτερικά της περίφραξης.",
        ],
    },
    {
        "title": "2. Λειτουργικά μέτρα για τη βιοπροφύλαξη",
        "items": [
            "Υγιεινή χεριών",
            "Ιματισμός & υποδήματα προσωπικού αποκλειστικά για την εκμετάλλευση και μιας χρήσης για τους επισκέπτες.",
            "Διαχωρισμός ζώων ανάλογα με την κατάσταση υγείας",
            "Ύπαρξη εντομοπαγίδων σε λειτουργία",
            "Ύπαρξη σταθμών μυοκτονίας σε λειτουργία",
            "Πρόγραμμα εντομοκτονίας - μυοκτονίας σε εφαρμογή",
            "Ορθή αποθήκευση ζωοτροφών και στρωμνής",
            "Ύπαρξη διαδικασίας ελέγχου του νερού",
        ],
    },
    {
        "title": "3. Πρόγραμμα επιτήρησης της υγείας των ζώων",
        "items": [
            "Ενημέρωση αρμόδιας αρχής επί υπόνοιας νοσήματος",
            "Τεκμηρίωση της προέλευσης των νεοεισερχόμενων ζώων.",
            "Τήρηση καραντίνας νεοεισερχόμενων ζώων (4 εβδομάδες)",
        ],
    },
    {
        "title": "4. Καθαριότητα και απολύμανση",
        "items": [
            "Περιβάλλων χώρος καθαρός.",
            "Καθαρός εξοπλισμός (τροφοδόχοι και υδροδόχοι).",
            "Χρήση εγκεκριμένων κατάλληλων απολυμαντικών (παραστατικά αγοράς)",
            "Πρόγραμμα καθαρισμών χώρων και εξοπλισμού – απολυμάνσεων από πιστοποιημένους φορείς",
            "Aπουσία άχρηστων αντικειμένων στο χώρο.",
            "Αντικατάσταση στρωμνής με γραπτή διαδικασία.",
        ],
    },
    {
        "title": "5. Τήρηση αρχείων",
        "items": [
            "Βιβλίο επισκεπτών–οχημάτων συμπληρωμένο (συμπληρωμένο/επικαιροποιημένο)",
            "Μητρώο εκμετάλλευσης (συμπληρωμένο/επικαιροποιημένο)",
            "Μητρώο φαρμακευτικής αγωγής (συμπληρωμένο/επικαιροποιημένο)",
            "Αρχείο ελέγχου αναλύσεων του νερού της εκμετάλλευσης.",
            "Αρχείο παραστατικών αγοράς ζωοτροφών",
            "Αρχείο ιδιοπαραγόμενων ζωοτροφών με υπολογισμό βάρους κατ' εκτίμηση.",
            "Βεβαίωση εκπαίδευσης.",
        ],
    },
]

CRITICAL_ITEM_INDICES = [
    {0, 1, 3, 8},
    {1, 2, 6},
    {0, 1, 2},
    {1, 2},
    {0, 1, 2, 4, 5, 6},
]

#----------------------------------------- ΣΤΑΘΕΡΕΣ ΒΑΡΩΝ, ΠΟΙΝΩΝ & ΧΡΩΜΑΤΩΝ ------------------------------------------#
ALERT_THRESHOLD_PCT = 30.0
CRITICAL_WEIGHT = 3.0
NORMAL_WEIGHT = 1.0

# Κλιμακούμενη ποινή ανά επιλογή -- κάθε στήλη ξεχωριστή βαρύτητα, με το
# "Μη Εφαρμόσιμο" να είναι η ΧΕΙΡΟΤΕΡΗ δυνατή απάντηση (μεγαλύτερη ποινή και από τη Σοβαρή
# Μη Συμμόρφωση). Χρησιμοποιείται παντού: οθόνη, Excel export, PDF export.
DEFICIENCY_BY_SCORE = {
    2: 0.0,            # ΣΥΜΜΟΡΦΩΣΗ
    1: 1.0 / 3.0,       # ΗΠΙΑ ΜΗ ΣΥΜΜΟΡΦΩΣΗ
    0: 2.0 / 3.0,       # ΣΟΒΑΡΗ ΜΗ ΣΥΜΜΟΡΦΩΣΗ
    -2: 1.0,            # ΜΗ ΕΦΑΡΜΟΣΙΜΟ -> μέγιστη ποινή
}

HEADER_BLUE = "#366092"
SECTION_BLUE = "#DCE6F1"
SECTION_BLUE_TXT = "#1F497D"
GOOD_GREEN_BG = "#C6EFCE"
GOOD_GREEN_TXT = "#006100"
WARN_YELLOW_BG = "#FFEB9C"
WARN_YELLOW_TXT = "#9C6500"
BAD_RED_BG = "#FFC7CE"
BAD_RED_TXT = "#9C0006"
ORANGE_BG = "#FCD5B4"
ORANGE_TXT = "#974706"
ALERT_BG = "#FF4D4D"
ALERT_TXT = "#FFFFFF"

# --- Παλέτα διεπαφής (καθαρά οπτικό, δεν επηρεάζει καμία λειτουργία) ---
APP_BG = "#EEF1F5"          # Απαλό φόντο γύρω από τις κάρτες
CARD_BG = "#FFFFFF"         # Φόντο καρτών/πλαισίων
CARD_BORDER = "#DDE3EA"     # Λεπτό περίγραμμα καρτών
ROW_ALT_BG = "#F6F8FA"      # Ζέβρα-ρίγα μη-κρίσιμων γραμμών
ACCENT_BLUE = "#2C5A8C"     # Hover/έντονο μπλε για κουμπιά


def deficiency_color(pct):
    if pct <= 30:
        return GOOD_GREEN_BG, GOOD_GREEN_TXT
    elif pct <= 50:
        return WARN_YELLOW_BG, WARN_YELLOW_TXT
    elif pct <= 90:
        return ORANGE_BG, ORANGE_TXT
    else:
        return BAD_RED_BG, BAD_RED_TXT


#--------------------------------------- ΚΥΡΙΑ ΕΦΑΡΜΟΓΗ (BiosafetyChecklistApp) ---------------------------------------#
class BiosafetyChecklistApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Λίστα Ελέγχου Βιοπροφύλαξης - Δείκτης Βαρύτητας")
        self.root.geometry("1180x880")
        self.root.minsize(980, 720)
        try:
            self.root.state("zoomed")  # Windows / macOS: μεγιστοποιημένο, με μπάρα τίτλου
        except tk.TclError:
            try:
                self.root.attributes("-zoomed", True)  # Linux (ορισμένοι window managers)
            except tk.TclError:
                sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
                self.root.geometry(f"{sw}x{sh}+0+0")

        self.unsaved_changes = False
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.configure(bg=APP_BG)

        self.section_vars = []
        self.section_total_labels = []
        self.section_alert_labels = []

        # --- ΜΕΤΑΒΛΗΤΕΣ ΙΣΤΟΡΙΚΟΥ & ΠΟΙΝΗΣ ---
        self.install_type_var = tk.StringVar(value="Πρόχειρο Κατάλυμα")
        self.install_type_var.trace_add("write", self._toggle_historical_frame)

        self.history_loaded = tk.BooleanVar(value=False)
        self.history_from_file = False
        self.history_months = tk.IntVar(value=0)
        self.history_penalty_weight = tk.DoubleVar(value=0.0)
        self.history_prev_score = tk.DoubleVar(value=0.0)

        # Χειροκίνητες μεταβλητές ιστορικού
        self.manual_score_var = tk.StringVar()
        self.manual_date_var = tk.StringVar()
        self.manual_score_var.trace_add("write", self._on_manual_history_change)
        self.manual_date_var.trace_add("write", self._on_manual_history_change)

        self.farm_name_var = tk.StringVar()
        self.inspector_var = tk.StringVar()
        self.date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))

        self.farm_name_var.trace_add("write", self._mark_unsaved)
        self.inspector_var.trace_add("write", self._mark_unsaved)
        self.date_var.trace_add("write", lambda *args: (self._on_manual_history_change(), self._recalculate()))

        self._build_ui()
        self._recalculate()

    def _mark_unsaved(self, *args):
        self.unsaved_changes = True

    def _add_hover(self, widget, normal_bg, hover_bg):
        """Καθαρά οπτικό εφέ hover (δεν αγγίζει command/λειτουργικότητα)."""
        widget.bind("<Enter>", lambda e: widget.config(bg=hover_bg))
        widget.bind("<Leave>", lambda e: widget.config(bg=normal_bg))

    def _build_ui(self):
        header = tk.Frame(self.root, bg=CARD_BG)
        header.pack(fill="x", padx=0, pady=0)

        title_bar = tk.Frame(header, bg=CARD_BG)
        title_bar.pack(fill="x")

        text_col = tk.Frame(title_bar, bg=CARD_BG)
        text_col.pack(side="left", fill="x", expand=True)

        tk.Label(
            text_col,
            text="ΥΠΟΥΡΓΕΙΟ ΑΓΡΟΤΙΚΗΣ ΑΝΑΠΤΥΞΗΣ ΚΑΙ ΤΡΟΦΙΜΩΝ - ΓΕΝΙΚΗ ΔΙΕΥΘΥΝΣΗ ΚΤΗΝΙΑΤΡΙΚΗΣ",
            font=("Segoe UI", 13, "bold"), fg=SECTION_BLUE_TXT, bg=CARD_BG, anchor="center",
        ).pack(fill="x", padx=18, pady=(12, 0))

        tk.Label(
            text_col,
            text="ΛΙΣΤΑ ΕΛΕΓΧΟΥ ΒΙΟΠΡΟΦΥΛΑΞΗΣ (ΚΑΤΑΓΡΑΦΗ ΣΥΜΜΟΡΦΩΣΗΣ)",
            font=("Segoe UI", 11, "bold"), fg="#333333", bg=CARD_BG, anchor="center",
        ).pack(fill="x", padx=18, pady=(3, 4))

        self._load_logo(title_bar)

        tk.Frame(header, bg=HEADER_BLUE, height=3).pack(fill="x")

        info_lbl = tk.Label(
            header,
            text="  Το ποσοστό (%) αποτυπώνει το ΔΕΙΚΤΗ ΒΑΡΥΤΗΤΑΣ (0% = Άριστα, 100% = Πλήρης Απουσία Βιοπροφύλαξης).   "
                 "  Τα μαρκαρισμένα με 🔺 / κίτρινο φόντο κριτήρια έχουν αυξημένη βαρύτητα. Όλα τα πεδία είναι υποχρεωτικά.",
            font=("Segoe UI", 8, "italic"), fg="#5A6472", bg="#F2F5F8", anchor="w", justify="left",
        )
        info_lbl.pack(fill="x", padx=0, pady=0, ipady=5, ipadx=18)
        # Το κείμενο δεν τυλίγει (wrap) από μόνο του χωρίς wraplength -- χωρίς αυτό κόβεται στο άκρο
        # του παραθύρου αντί να πάει σε 2η γραμμή. Το ενημερώνουμε δυναμικά ώστε να προσαρμόζεται
        # στο πλάτος του παραθύρου (μεγιστοποιημένο, restored, κ.λπ.) και να φαίνεται πάντα ολόκληρο.
        info_lbl.bind("<Configure>", lambda e: info_lbl.config(wraplength=max(300, e.width - 36)))

        # --- ΠΛΑΙΣΙΟ: Τύπος Εγκατάστασης & Ιστορικό ---
        self.history_frame = tk.LabelFrame(
            self.root, text=" Τύπος Εγκατάστασης & Ιστορικό Ελέγχων ",
            font=("Segoe UI", 9, "bold"), bg=CARD_BG, fg=SECTION_BLUE_TXT, padx=12, pady=8, relief="flat",
            bd=1, highlightbackground=CARD_BORDER, highlightthickness=1
        )
        self.history_frame.pack(fill="x", padx=14, pady=(10, 6))

        type_row = tk.Frame(self.history_frame, bg=CARD_BG)
        type_row.pack(fill="x", pady=(0, 6))

        tk.Label(type_row, text="Επιλογή Τύπου:", font=("Segoe UI", 9, "bold"), bg=CARD_BG, fg="#333333").pack(
            side="left", padx=(0, 10))
        tk.Radiobutton(type_row, text="Πρόχειρο Κατάλυμα", variable=self.install_type_var,
                       value="Πρόχειρο Κατάλυμα", bg=CARD_BG, font=("Segoe UI", 9),
                       selectcolor="white", cursor="hand2", activebackground=CARD_BG).pack(side="left", padx=10)
        tk.Radiobutton(type_row, text="Κτηνοτροφική Εγκατάσταση", variable=self.install_type_var,
                       value="Κτηνοτροφική Εγκατάσταση", bg=CARD_BG, font=("Segoe UI", 9),
                       selectcolor="white", cursor="hand2", activebackground=CARD_BG).pack(side="left", padx=10)

        self.excel_load_container = tk.Frame(self.history_frame, bg=CARD_BG)

        # Row 1: Excel Load Button & Status + Manual Inputs Side-by-Side Guide
        row1 = tk.Frame(self.excel_load_container, bg=CARD_BG)
        row1.pack(fill="x", pady=2)

        load_btn = tk.Button(
            row1, text="📁  Φόρτωση προηγούμενων Excel", command=self.load_historical_files,
            bg=HEADER_BLUE, fg="white", font=("Segoe UI", 9, "bold"), cursor="hand2", padx=12, pady=6,
            relief="flat", bd=0, activebackground=ACCENT_BLUE, activeforeground="white"
        )
        load_btn.pack(side="left")
        self._add_hover(load_btn, HEADER_BLUE, ACCENT_BLUE)

        self.history_status_lbl = tk.Label(row1, text="Δεν έχει φορτωθεί αρχείο ή εισάγετε τα στοιχεία χειροκίνητα", bg=CARD_BG,
                                           font=("Segoe UI", 9, "italic"), fg="#666666")
        self.history_status_lbl.pack(side="left", padx=12)

        # Χειροκίνητη Εισαγωγή δίπλα σαν οδηγία (χωρίς την παλιά επικεφαλίδα)
        manual_frame = tk.Frame(self.excel_load_container, bg=CARD_BG)
        manual_frame.pack(fill="x", pady=(8, 3))

        tk.Label(manual_frame, text="Προηγ. Δείκτης (%):", bg=CARD_BG, fg="#555555", font=("Segoe UI", 8)).pack(
            side="left")
        tk.Entry(manual_frame, textvariable=self.manual_score_var, width=8, font=("Segoe UI", 8),
                 relief="solid", bd=1, highlightthickness=1, highlightcolor=HEADER_BLUE,
                 highlightbackground=CARD_BORDER).pack(side="left", padx=(4, 12))

        tk.Label(manual_frame, text="Ημερομηνία (ΗΗ/ΜΜ/ΕΕΕΕ):", bg=CARD_BG, fg="#555555", font=("Segoe UI", 8)).pack(
            side="left")
        tk.Entry(manual_frame, textvariable=self.manual_date_var, width=12, font=("Segoe UI", 8),
                 relief="solid", bd=1, highlightthickness=1, highlightcolor=HEADER_BLUE,
                 highlightbackground=CARD_BORDER).pack(side="left", padx=(4, 12))

        reset_hist_btn = tk.Button(manual_frame, text="Καθαρισμός Ιστορικού", command=self._reset_historical_data,
                                   bg="#8A8A8A", fg="white", font=("Segoe UI", 7, "bold"), padx=7, pady=3,
                                   relief="flat", bd=0, cursor="hand2", activebackground="#6E6E6E",
                                   activeforeground="white")
        reset_hist_btn.pack(side="left")
        self._add_hover(reset_hist_btn, "#8A8A8A", "#6E6E6E")

        # Row 3: Legend of Penalty Weights
        row3 = tk.Frame(self.excel_load_container, bg=CARD_BG)
        row3.pack(fill="x", pady=(6, 0))

        tk.Label(row3, text="Ποινή Ιστορικού (μόνο αν προηγ. ποσοστό ≥30%):",
                 font=("Segoe UI", 8, "bold"), bg=CARD_BG, fg="#555555").pack(side="left", padx=(0, 6))

        tk.Label(row3, text="0-6 μήνες: Ποινή ", bg="#FFC7CE", fg="#9C0006", font=("Segoe UI", 8, "bold"), padx=7,
                 pady=2).pack(side="left", padx=2)
        tk.Label(row3, text="6-12 μήνες: Ποινή", bg="#FFEB9C", fg="#9C6500", font=("Segoe UI", 8, "bold"), padx=7,
                 pady=2).pack(side="left", padx=2)
        tk.Label(row3, text="12-24 μήνες: Ποινή ", bg="#C6EFCE", fg="#006100", font=("Segoe UI", 8, "bold"), padx=7,
                 pady=2).pack(side="left", padx=2)
        tk.Label(row3, text="24+ μήνες Χωρίς Ποινή", bg="#E8E8E8", fg="#555555", font=("Segoe UI", 8, "bold"),
                 padx=7, pady=2).pack(side="left", padx=2)

        info = tk.Frame(self.root, bg=CARD_BG, relief="flat", bd=1, highlightbackground=CARD_BORDER,
                        highlightthickness=1)
        info.pack(fill="x", padx=14, pady=(0, 8))

        info_inner = tk.Frame(info, bg=CARD_BG)
        info_inner.pack(fill="x", padx=10, pady=9)

        entry_style = dict(relief="solid", bd=1, highlightthickness=1,
                          highlightcolor=HEADER_BLUE, highlightbackground=CARD_BORDER, font=("Segoe UI", 9))

        tk.Label(info_inner, text="Εκμετάλλευση - Κωδικός*:", bg=CARD_BG, fg="#333333", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="w", padx=(0, 6))
        tk.Entry(info_inner, textvariable=self.farm_name_var, width=30, **entry_style).grid(row=0, column=1,
                                                                                             sticky="w", padx=(0, 22))

        tk.Label(info_inner, text="Ελεγκτής/τρια*:", bg=CARD_BG, fg="#333333", font=("Segoe UI", 9)).grid(
            row=0, column=2, sticky="w", padx=(0, 6))
        tk.Entry(info_inner, textvariable=self.inspector_var, width=25, **entry_style).grid(row=0, column=3,
                                                                                             sticky="w", padx=(0, 22))

        tk.Label(info_inner, text="Ημερομηνία (ΗΗ/ΜΜ/ΕΕΕΕ)*:", bg=CARD_BG, fg="#333333", font=("Segoe UI", 9)).grid(
            row=0, column=4, sticky="w", padx=(0, 6))
        tk.Entry(info_inner, textvariable=self.date_var, width=12, **entry_style).grid(row=0, column=5, sticky="w")

        container = tk.Frame(self.root, bg=APP_BG)
        container.pack(fill="both", expand=True, padx=14, pady=(0, 8))

        canvas = tk.Canvas(container, borderwidth=0, highlightthickness=0, bg=CARD_BG)
        vscroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.scroll_frame = tk.Frame(canvas, bg=CARD_BG)

        self.scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas_window = canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")

        def _configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", _configure_canvas)
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        def _on_mousewheel(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel)
        canvas.bind_all("<Button-5>", _on_mousewheel)

        self._build_table_header(self.scroll_frame)

        for s_idx, section in enumerate(SECTIONS):
            self._build_section(self.scroll_frame, s_idx, section)

        footer = tk.Frame(self.root, bg=CARD_BG, relief="flat", bd=0,
                          highlightbackground=CARD_BORDER, highlightthickness=1)
        footer.pack(fill="x", padx=14, pady=(0, 12))

        totals_frame = tk.Frame(footer, bg=CARD_BG)
        totals_frame.pack(fill="x", padx=12, pady=10)

        self.progress_bar = ttk.Progressbar(totals_frame, orient="horizontal", length=200, mode="determinate",
                                            style="Green.Horizontal.TProgressbar")
        self.progress_bar.pack(side="left", padx=(10, 6))

        self.progress_lbl = tk.Label(totals_frame, text="Απαντημένα: 0/34", font=("Segoe UI", 10), bg=CARD_BG,
                                     fg="#333333")
        self.progress_lbl.pack(side="left", padx=(0, 10))

        self.pct_label = tk.Label(totals_frame, text="0.0%", font=("Segoe UI", 16, "bold"),
                                  bg=GOOD_GREEN_BG, fg=GOOD_GREEN_TXT, width=8, relief="flat", bd=0,
                                  anchor="center", padx=6, pady=4)
        self.pct_label.pack(side="right", padx=(0, 10))

        tk.Label(totals_frame, text="ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ ΒΑΡΥΤΗΤΑΣ ΜΗ ΣΥΜΜΟΡΦΩΣΗΣ [0-1] norm%:",
                font=("Segoe UI", 11, "bold"), bg=CARD_BG, fg="#333333", anchor="e").pack(side="right", padx=(0, 10))

        # --- ΠΛΑΙΣΙΟ: Επίδραση Προηγούμενου Ελέγχου (Ποινή) ---
        self.hist_penalty_frame = tk.Frame(footer, bg="#E9EBEE", relief="flat", bd=0)
        self.hist_penalty_frame.pack(fill="x", padx=12, pady=(0, 10))

        self.hist_penalty_lbl = tk.Label(
            self.hist_penalty_frame,
            text="Προηγούμενη Συμμόρφωση: Απενεργοποιημένο (Προστιθέμενη Ποινή: 0.0%)",
            font=("Segoe UI", 10, "italic"), bg="#E9EBEE", fg="#666666"
        )
        self.hist_penalty_lbl.pack(padx=10, pady=(7, 0))

        self.history_files_lbl = tk.Label(
            self.hist_penalty_frame, text="", font=("Segoe UI", 9), bg="#E9EBEE", fg="#8A8F98",
            justify="left", anchor="w"
        )
        self.history_files_lbl.pack(padx=10, pady=(0, 6), fill="x")

        btns = tk.Frame(footer, bg=CARD_BG)
        btns.pack(fill="x", padx=12, pady=(0, 12))

        clear_btn = tk.Button(btns, text="🗑  Καθαρισμός φόρμας", command=self.clear_form,
                              bg="#E4E7EB", fg="#333333", font=("Segoe UI", 9), padx=10, pady=6, relief="flat",
                              bd=0, cursor="hand2", activebackground="#D2D7DD", activeforeground="#333333")
        clear_btn.pack(side="left", padx=4)
        self._add_hover(clear_btn, "#E4E7EB", "#D2D7DD")

        pdf_btn = tk.Button(btns, text="📄  Εξαγωγή σε PDF (.pdf)", command=self.export_to_pdf,
                            bg=BAD_RED_TXT, fg="white", font=("Segoe UI", 9, "bold"), padx=12, pady=6,
                            relief="flat", bd=0, cursor="hand2", activebackground="#7A0005",
                            activeforeground="white")
        pdf_btn.pack(side="right", padx=4)
        self._add_hover(pdf_btn, BAD_RED_TXT, "#7A0005")

        excel_btn = tk.Button(btns, text="📊  Εξαγωγή σε Excel (.xlsx)", command=self.export_to_excel,
                              bg=HEADER_BLUE, fg="white", font=("Segoe UI", 9, "bold"), padx=12, pady=6,
                              relief="flat", bd=0, cursor="hand2", activebackground=ACCENT_BLUE,
                              activeforeground="white")
        excel_btn.pack(side="right", padx=4)
        self._add_hover(excel_btn, HEADER_BLUE, ACCENT_BLUE)

        legend_frame = tk.Frame(footer, bg=CARD_BG)
        legend_frame.pack(fill="x", padx=12, pady=(0, 6))

        tk.Label(legend_frame, text="Υπόμνημα Κατηγοριών:", font=("Segoe UI", 7, "bold"),
                 bg=CARD_BG, fg="#777777").pack(side="left", padx=(0, 6))
        for text, bg, fg in [
            ("0-30% Συμμόρφωση", GOOD_GREEN_BG, GOOD_GREEN_TXT),
            ("30-50% Ήπια Μη Συμμόρφωση", WARN_YELLOW_BG, WARN_YELLOW_TXT),
            ("50-90% Σοβαρή Μη Συμμόρφωση", ORANGE_BG, ORANGE_TXT),
            ("90-100% Μη Εφαρμόσιμο", BAD_RED_BG, BAD_RED_TXT),
        ]:
            tk.Label(legend_frame, text=text, font=("Segoe UI", 7, "bold"), bg=bg, fg=fg,
                     padx=6, pady=2).pack(side="left", padx=(0, 4))

        tk.Label(footer, text="Created by © 2026 Aggelos A. Kaptanis", font=("Segoe UI", 7, "italic"),
                 fg="#AAAAAA", bg=CARD_BG, anchor="w").pack(fill="x", padx=10, pady=(0, 4))

        self._toggle_historical_frame()

    # --- ΛΟΓΙΚΗ ΙΣΤΟΡΙΚΟΥ & EXCEL ---
    #------------------------------------- ΙΣΤΟΡΙΚΟ ΕΛΕΓΧΩΝ & ΥΠΟΛΟΓΙΣΜΟΣ ΠΟΙΝΗΣ --------------------------------------#
    def _toggle_historical_frame(self, *args):
        if self.install_type_var.get() == "Κτηνοτροφική Εγκατάσταση":
            self.excel_load_container.pack(fill="x", expand=True)
        else:
            self.excel_load_container.pack_forget()
            self._reset_historical_data()

    FARM_CODE_RE = re.compile(r'[A-Za-zΑ-Ωα-ω]{2}\s*-?\s*\d{4,}')

    # Ελληνικά πλήκτρα -> λατινικά, με βάση τη ΘΕΣΗ στο πληκτρολόγιο (όχι οπτική ομοιότητα) --
    # π.χ. αν κάποιος έγραψε "EL" με κατά λάθος ενεργό ελληνικό πληκτρολόγιο βγήκε "ΕΛ". Χωρίς αυτή
    # την κανονικοποίηση, οι δύο μορφές του ίδιου κωδικού θα φαίνονταν σαν διαφορετικός κωδικός.
    _GREEK_KB_TO_LATIN = str.maketrans({
        "Α": "A", "Β": "B", "Ψ": "C", "Δ": "D", "Ε": "E", "Φ": "F", "Γ": "G", "Η": "H",
        "Ι": "I", "Ξ": "J", "Κ": "K", "Λ": "L", "Μ": "M", "Ν": "N", "Ο": "O", "Π": "P",
        "Ρ": "R", "Σ": "S", "Τ": "T", "Θ": "U", "Ω": "V", "Χ": "X", "Υ": "Y", "Ζ": "Z",
    })

    def _extract_farm_code(self, text):
        """Εξάγει τον κωδικό εκμετάλλευσης (π.χ. EL434343) από ελεύθερο κείμενο -- αγνοεί τα υπόλοιπα."""
        if not text:
            return None
        m = self.FARM_CODE_RE.search(str(text).upper())
        if not m:
            return None
        code = m.group(0).replace(" ", "").replace("-", "")
        return code.translate(self._GREEK_KB_TO_LATIN)

    def _compute_history_penalty(self):
        """Ποινή ιστορικού:
        - Προηγ. ποσοστό μη συμμόρφωσης < 30% -> καμία ποινή.
        - 24+ μήνες -> παραγραφή, καμία ποινή.
        - Αλλιώς: 0-6 μήνες -> Ποσοστό/2 | 6-12 μήνες -> Ποσοστό/3 | 12-24 μήνες -> Ποσοστό/4.
          (π.χ. προηγ. 50%, 0-6 μήνες -> 50/2 = +25.00%)
        Επιστρέφει (penalty_pct, applies, note) -- penalty_pct σε μονάδες ποσοστού (π.χ. 25.00 = +25.00%)."""
        if not self.history_loaded.get():
            return 0.0, False, ""

        prev_pct = self.history_prev_score.get() * 100.0
        months = self.history_months.get()

        if prev_pct < 30.0:
            return 0.0, False, f"Προηγ. ποσοστό {prev_pct:.2f}% (κάτω από το όριο 30%)"
        if months > 24:
            return 0.0, False, f"{months} μήνες (Παραγραφή)"

        if months <= 6:
            return prev_pct / 2.0, True, "0-6 μήνες"
        elif months <= 12:
            return prev_pct / 3.0, True, "6-12 μήνες"
        else:
            return prev_pct / 4.0, True, "12-24 μήνες"

    def _reset_historical_data(self):
        self.manual_score_var.set("")
        self.manual_date_var.set("")
        self.history_loaded.set(False)
        self.history_from_file = False
        self.history_months.set(0)
        self.history_penalty_weight.set(0.0)
        self.history_prev_score.set(0.0)
        self.history_status_lbl.config(text="Δεν έχει φορτωθεί αρχείο ή εισάγετε τα στοιχεία χειροκίνητα", fg="#555555")

        self.hist_penalty_frame.config(bg="#DDDDDD")
        self.hist_penalty_lbl.config(
            text="Προηγούμενη Συμμόρφωση: Απενεργοποιημένο (Προστιθέμενη Ποινή: 0.0%)",
            bg="#DDDDDD", fg="#666666", font=("Segoe UI", 10, "italic")
        )
        self.history_files_lbl.config(text="")
        self._recalculate()

    def _on_manual_history_change(self, *args):
        score_str = self.manual_score_var.get().strip()
        date_str = self.manual_date_var.get().strip()

        if not score_str or not date_str:
            if self.history_loaded.get() and not self.history_from_file:
                self.history_loaded.set(False)
                self.history_penalty_weight.set(0.0)
                self.history_prev_score.set(0.0)
                self.history_status_lbl.config(text="Δεν έχει φορτωθεί αρχείο ή να βάλτε τα στοιχεία χειροκίνητα", fg="#555555")
                self.history_files_lbl.config(text="")
                self._recalculate()
            return

        try:
            score_val = float(score_str.replace(',', '.'))
            if score_val > 1.0:
                score_val /= 100.0

            prev_date = datetime.strptime(date_str, "%d/%m/%Y")
            current_date_str = self.date_var.get().strip()
            current_date = datetime.strptime(current_date_str, "%d/%m/%Y")

            days_diff = (current_date - prev_date).days
            if days_diff < 0:
                self.history_status_lbl.config(text="⚠️ Η ημερομηνία προηγ. ελέγχου είναι μεταγενέστερη!", fg="#9C0006")
                return

            months_diff = days_diff / 30.44

            self.history_loaded.set(True)
            self.history_from_file = False
            self.history_months.set(int(months_diff))
            self.history_prev_score.set(score_val)
            self.history_files_lbl.config(text="")

            penalty_pct, applies, note = self._compute_history_penalty()
            self.history_penalty_weight.set(1.0 if applies else 0.0)

            if applies:
                self.history_status_lbl.config(
                    text=f"Χειροκίνητη εισαγωγή: {score_val * 100:.2f}% | {note} | Ποινή: +{penalty_pct:.2f}%",
                    fg="#006100")
            else:
                self.history_status_lbl.config(
                    text=f"Χειροκίνητη εισαγωγή: {score_val * 100:.2f}% | {note}", fg="#555555")
            self._recalculate()

        except Exception:
            pass

    def load_historical_files(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Λείπει βιβλιοθήκη", "Το openpyxl απαιτείται για την ανάγνωση αρχείων.")
            return

        filepaths = filedialog.askopenfilenames(
            title="Επιλογή προηγούμενων ελέγχων (Excel)",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not filepaths:
            return

        current_date_str = self.date_var.get().strip()
        try:
            current_date = datetime.strptime(current_date_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Λάθος", "Η τρέχουσα ημερομηνία δεν έχει σωστή μορφή (ΗΗ/ΜΜ/ΕΕΕΕ).")
            return

        results = []

        for path in filepaths:
            try:
                filename = os.path.basename(path)
                wb = openpyxl.load_workbook(path, data_only=True)

                file_date = None
                score = None
                farm_code = None

                # --- Προτιμώμενη πηγή: κρυφό φύλλο "Meta" (καθαρές τιμές, όχι μορφοποιημένο κείμενο) ---
                if "Meta" in wb.sheetnames:
                    meta = wb["Meta"]
                    farm_code = self._extract_farm_code(str(meta["B1"].value or ""))
                    date_raw = meta["B2"].value
                    if isinstance(date_raw, str):
                        try:
                            file_date = datetime.strptime(date_raw.strip(), "%d/%m/%Y")
                        except ValueError:
                            file_date = None
                    elif hasattr(date_raw, "year"):
                        file_date = datetime(date_raw.year, date_raw.month, date_raw.day)
                    pct_raw = meta["B3"].value
                    if isinstance(pct_raw, (int, float)):
                        score = float(pct_raw)

                ws = wb.active

                # --- Fallback ημερομηνίας: όνομα αρχείου, αλλιώς κείμενο B4 (παλιότερα αρχεία χωρίς Meta) ---
                if file_date is None:
                    match = re.search(r'(\d{2}-\d{2}-\d{4})\.xlsx$', filename)
                    if match:
                        file_date = datetime.strptime(match.group(1), "%d-%m-%Y")
                    else:
                        b4_val = str(ws["B4"].value or "")
                        date_match = re.search(r'Ημερομηνία:\s*(\d{2}/\d{2}/\d{4})', b4_val)
                        if date_match:
                            file_date = datetime.strptime(date_match.group(1), "%d/%m/%Y")

                if file_date is None:
                    continue

                # --- Fallback κωδικού εκμετάλλευσης: κείμενο B4 ---
                if farm_code is None:
                    farm_code = self._extract_farm_code(str(ws["B4"].value or ""))

                # --- Fallback ποσοστού: σάρωση για "ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ" (στήλη G τωρινή διάταξη, F παλιά αρχεία) ---
                if score is None:
                    for row in range(ws.max_row, 1, -1):
                        cell_val = str(ws.cell(row=row, column=2).value or "").upper()
                        if ("ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ" in cell_val or "ΔΕΙΚΤΗΣ ΒΑΡΥΤΗΤΑΣ" in cell_val
                                or "ΣΥΝΟΛΟ ΣΥΜΜΟΡΦΩΣΗΣ" in cell_val):
                            for col in (7, 6):
                                val = ws.cell(row=row, column=col).value
                                if isinstance(val, (int, float)):
                                    score = float(val)
                                    break
                                if isinstance(val, str):
                                    try:
                                        score = float(val.replace('%', '').replace(',', '.')) / 100.0
                                        break
                                    except ValueError:
                                        pass
                            if score is not None:
                                break

                    # Εναλλακτική σάρωση σε περίπτωση που ο δείκτης βρίσκεται σε άλλη θέση
                    if score is None:
                        for row in range(1, ws.max_row + 1):
                            for col in range(1, ws.max_column + 1):
                                cell_txt = str(ws.cell(row=row, column=col).value or "").upper()
                                if "ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ" in cell_txt or "ΔΕΙΚΤΗΣ ΒΑΡΥΤΗΤΑΣ" in cell_txt:
                                    val = ws.cell(row=row, column=col + 4).value if col + 4 <= ws.max_column else None
                                    if isinstance(val, (int, float)):
                                        score = float(val)
                                        break
                            if score is not None:
                                break

                if score is None:
                    continue

                days_diff = (current_date - file_date).days
                if days_diff < 0:
                    continue

                if score > 1.0:
                    score /= 100.0

                results.append({
                    "filename": filename, "months_diff": days_diff / 30.44,
                    "score": score, "farm_code": farm_code, "file_date": file_date,
                    "filename_code": self._extract_farm_code(filename),
                })

            except Exception as e:
                print(f"Σφάλμα ανάγνωσης {path}: {e}")
                continue

        if not results:
            messagebox.showwarning("Αποτυχία", "Δεν βρέθηκαν έγκυρα παλαιότερα αρχεία.")
            return

        # --- Αμυντικός έλεγχος κωδικού εκμετάλλευσης, με βάση τον κωδικό μέσα στο ΟΝΟΜΑ κάθε
        # αρχείου (π.χ. Checklist_Biosecurity_EL4344_08-08-2026.xlsx) -- μία μόνο πηγή αλήθειας. ---

        # 1) Όλα τα επιλεγμένα αρχεία πρέπει να έχουν τον ίδιο κωδικό εκμετάλλευσης μεταξύ τους.
        found_codes = {r["filename_code"] for r in results if r["filename_code"]}
        if len(found_codes) > 1:
            messagebox.showerror(
                "Διαφορετικοί κωδικοί εκμετάλλευσης",
                "Τα αρχεία που επέλεξες δεν έχουν όλα τον ίδιο κωδικό εκμετάλλευσης:\n" +
                ", ".join(sorted(found_codes)) + "\n\nΕπέλεξε μόνο αρχεία excel με ίδιο κωδικό εκμετάλλευσης."
            )
            return

        loaded_code = next(iter(found_codes), None)

        # 2) Αν έχεις ήδη γράψει κωδικό στη φόρμα, πρέπει να ταιριάζει με των αρχείων -- αν η φόρμα
        # είναι άδεια, τη συμπληρώνουμε αυτόματα (μόνο εφόσον τα αρχεία συμφωνούν μεταξύ τους, βήμα 1).
        current_code = self._extract_farm_code(self.farm_name_var.get())
        if loaded_code:
            if not current_code:
                self.farm_name_var.set(loaded_code)
            elif current_code != loaded_code:
                messagebox.showerror(
                    "Διαφορετικός κωδικός εκμετάλλευσης",
                    f"Ο κωδικός στη φόρμα ({current_code}) δεν ταιριάζει με τον κωδικό των αρχείων "
                    f"που φόρτωσες ({loaded_code})."
                )
                return

        # Prioritizes smallest months_diff; if tied, picks highest score
        closest = min(results, key=lambda r: (r["months_diff"], -r["score"]))

        self.history_loaded.set(True)
        self.history_from_file = True
        self.history_months.set(int(closest["months_diff"]))
        self.history_prev_score.set(closest["score"])

        penalty_pct, applies, note = self._compute_history_penalty()
        self.history_penalty_weight.set(1.0 if applies else 0.0)

        if applies:
            self.history_status_lbl.config(
                text=f"Αρχείο: {closest['filename']} | {note} | Ποινή: +{penalty_pct:.2f}%", fg="#006100")
        else:
            self.history_status_lbl.config(text=f"Αρχείο: {closest['filename']} | {note}", fg="#555555")

        # Συμπληρώνουμε ΚΑΙ τα δύο πεδία (ποσοστό + ημερομηνία) -- προσωρινά χωρίς trace ώστε να μην
        # ενεργοποιηθεί η χειροκίνητη-εισαγωγή λογική (θα ξανά-υπολόγιζε και θα έσβηνε το
        # history_from_file=True παραπάνω). Και τα δύο .set() πρέπει να γίνουν ΜΕΣΑ σε αυτό το μπλοκ.
        for var in (self.manual_score_var, self.manual_date_var):
            for mode, cbname in var.trace_info():
                var.trace_remove(mode, cbname)
        self.manual_score_var.set(f"{closest['score'] * 100:.2f}")
        self.manual_date_var.set(closest["file_date"].strftime("%d/%m/%Y"))
        self.manual_score_var.trace_add("write", self._on_manual_history_change)
        self.manual_date_var.trace_add("write", self._on_manual_history_change)

        self._update_history_files_summary(results, closest["filename"])
        self._recalculate()

    def _update_history_files_summary(self, results, selected_filename):
        """Μικρή περιοχή κάτω από τον κύριο δείκτη ποινής: όλα τα φορτωμένα αρχεία (ημερομηνία,
        ποσοστό), ένα κάτω από το άλλο, με ✓ στο επιλεγμένο."""
        if not results:
            self.history_files_lbl.config(text="")
            return
        lines = ["Φορτωμένα αρχεία:"]
        for r in sorted(results, key=lambda r: r["months_diff"]):
            mark = "✓ " if r["filename"] == selected_filename else "   "
            date_str = r["file_date"].strftime("%d/%m/%Y")
            lines.append(f"{mark}{r['filename']}  ({date_str}, {r['score'] * 100:.2f}%)")
        self.history_files_lbl.config(text="\n".join(lines))

    # --- UI HELPERS ---
    #------------------------------ ΒΟΗΘΗΤΙΚΑ UI ΣΤΟΙΧΕΙΑ (ΛΟΓΟΤΥΠΟ, ΠΙΝΑΚΑΣ ΚΡΙΤΗΡΙΩΝ) -------------------------------#
    def _load_logo(self, parent):
        logo_path = resource_path(LOGO_FILENAME)
        try:
            if not os.path.isfile(logo_path):
                return
            img = Image.open(logo_path)
            max_h = 70
            ratio = max_h / img.height
            new_size = (max(1, int(img.width * ratio)), max_h)
            img = img.resize(new_size, Image.LANCZOS)
            self._logo_imgtk = ImageTk.PhotoImage(img)
            logo_lbl = tk.Label(parent, image=self._logo_imgtk, bg=CARD_BG)
            logo_lbl.pack(side="right", padx=14, pady=8)
        except Exception:
            pass

    def _build_table_header(self, parent):
        row = tk.Frame(parent, bg=HEADER_BLUE)
        row.pack(fill="x", pady=(0, 0))

        # Καθαρά ονόματα χωρίς περιττά κενά
        headers = [
            "Α/Α", "Κριτήριο Ελέγχου", "ΣΥΜΜΟΡΦΩΣΗ",
            "ΗΠΙΑ ΜΗ ΣΥΜΜΟΡΦΩΣΗ", "ΣΟΒΑΡΗ ΜΗ ΣΥΜΜΟΡΦΩΣΗ", "ΜΗ ΕΦΑΡΜΟΣΙΜΟ"
        ]

        # Αρκετό width για να μην κόβονται τα γράμματα
        widths = [5, 70, 16, 20, 24, 16]
        aligns = ["center", "center", "e", "e", "e", "e"]

        # ΕΔΩ ΕΙΝΑΙ ΤΟ ΜΥΣΤΙΚΟ:
        # Βάζουμε 0 στις 3 επίμαχες στήλες ώστε το κείμενο να "κολλήσει" τέρμα δεξιά,
        # χωρίς να το σταματάει το αόρατο περιθώριο. (Αν θες να πάνε λίγο αριστερά, το κάνεις 2 ή 5).
        pad_x_list = [14, 14, 7, 7, 7, 14]

        for i, (h, w, al, px) in enumerate(zip(headers, widths, aligns, pad_x_list)):
            tk.Label(row, text=h, bg=HEADER_BLUE, fg="white", font=("Segoe UI", 9, "bold"),
                     width=w, anchor=al,
                     justify=("right" if al == "e" else ("left" if al == "w" else "center")),
                     padx=px, pady=13).grid(row=0, column=i, sticky="nsew")

        row.grid_columnconfigure(1, weight=1)
        tk.Frame(parent, bg=ACCENT_BLUE, height=2).pack(fill="x")

    def _build_section(self, parent, s_idx, section):
        title_row = tk.Frame(parent, bg=SECTION_BLUE)
        title_row.pack(fill="x", pady=(14, 0))
        tk.Label(title_row, text=section["title"], bg=SECTION_BLUE, fg=SECTION_BLUE_TXT,
                 font=("Segoe UI", 10, "bold"), anchor="w", padx=12, pady=8).pack(fill="x")

        critical_idx = CRITICAL_ITEM_INDICES[s_idx]
        vars_for_section = []
        for item_idx, item_text in enumerate(section["items"]):
            is_critical = item_idx in critical_idx
            row_bg = "#FFF6E5" if is_critical else (CARD_BG if item_idx % 2 == 0 else ROW_ALT_BG)

            item_row = tk.Frame(parent, bg=row_bg)
            item_row.pack(fill="x")

            var = tk.IntVar(value=-1)
            var.trace_add("write", self._on_radio_change)
            vars_for_section.append(var)

            marker = "🔺 " if is_critical else ""
            tk.Label(item_row, text="", width=5, bg=row_bg).grid(row=0, column=0, sticky="w")
            tk.Label(item_row, text=marker + item_text, bg=row_bg, fg="#2A2A2A", anchor="w", justify="left",
                     wraplength=560, padx=4, pady=6,
                     font=("Segoe UI", 9, "bold" if is_critical else "normal")).grid(row=0, column=1, sticky="w")

            item_row.grid_columnconfigure(1, weight=1)

            rb_frame = tk.Frame(item_row, bg=row_bg)
            rb_frame.grid(row=0, column=2, columnspan=3, sticky="e", padx=10)

            tk.Radiobutton(rb_frame, text="Συμμόρφωση", variable=var, value=2, bg=row_bg,
                          font=("Segoe UI", 9), selectcolor="white", cursor="hand2",
                          activebackground=row_bg).pack(side="left", padx=6)
            tk.Radiobutton(rb_frame, text="Ήπια Μη Συμμόρφωση", variable=var, value=1, bg=row_bg,
                          font=("Segoe UI", 9), selectcolor="white", cursor="hand2",
                          activebackground=row_bg).pack(side="left", padx=6)
            tk.Radiobutton(rb_frame, text="Σοβαρή Μη Συμμόρφωση", variable=var, value=0, bg=row_bg,
                          font=("Segoe UI", 9), selectcolor="white", cursor="hand2",
                          activebackground=row_bg).pack(side="left", padx=6)
            tk.Radiobutton(rb_frame, text="Μη Εφαρμόσιμο", variable=var, value=-2, bg=row_bg,
                          font=("Segoe UI", 9), selectcolor="white", cursor="hand2",
                          activebackground=row_bg).pack(side="left", padx=6)

            tk.Frame(parent, bg=CARD_BORDER, height=1).pack(fill="x")

        self.section_vars.append(vars_for_section)

        total_row = tk.Frame(parent, bg="#E8E8E8")
        total_row.pack(fill="x", pady=(0, 0))

        lbl = tk.Label(total_row, text="—", bg="#E8E8E8", fg="#333333", font=("Segoe UI", 10, "bold"), width=8,
                       anchor="center")
        lbl.pack(side="right", padx=(0, 10))

        tk.Label(total_row, text="Σύνολο Συμμόρφωσης Ενότητας:", bg="#E8E8E8", fg="#333333",
                 font=("Segoe UI", 9, "bold"), anchor="e", padx=8, pady=6).pack(side="right", fill="x", expand=True)

        self.section_total_labels.append(lbl)

        if critical_idx:
            alert_row = tk.Frame(parent, bg="#F5F5F5")
            alert_row.pack(fill="x", pady=(0, 0))

            pct_value_lbl = tk.Label(alert_row, text="—", bg="#F5F5F5", fg="#555555", font=("Segoe UI", 10, "bold"),
                                     width=8, anchor="center")
            pct_value_lbl.pack(side="right", padx=(0, 10))

            alert_lbl = tk.Label(alert_row, text="", font=("Segoe UI", 9, "bold"), bg="#F5F5F5", anchor="e", padx=8,
                                 pady=4, justify="right")
            alert_lbl.pack(side="right", fill="x", expand=True)

            pct_lbl = tk.Label(alert_row, text="Δείκτης Συναγερμού Κρίσιμων:", bg="#F5F5F5", fg="#555555",
                               font=("Segoe UI", 8, "bold"), anchor="w", padx=8, pady=4)
            pct_lbl.pack(side="left")
            self.section_alert_labels.append((alert_row, pct_value_lbl, alert_lbl))
        else:
            self.section_alert_labels.append(None)

    #-------------------------------------------- ΥΠΟΛΟΓΙΣΜΟΙ ΒΑΘΜΟΛΟΓΙΑΣ ---------------------------------------------#
    def _on_radio_change(self, *args):
        self.unsaved_changes = True
        self._recalculate()

    def _section_deficiency_ratio(self, s_idx):
        critical_idx = CRITICAL_ITEM_INDICES[s_idx]
        weighted_deficiency = 0.0
        weighted_total = 0.0
        any_answered = False

        for i, v in enumerate(self.section_vars[s_idx]):
            score = v.get()
            if score == -1:
                continue
            any_answered = True
            weight = CRITICAL_WEIGHT if i in critical_idx else NORMAL_WEIGHT
            deficiency = DEFICIENCY_BY_SCORE[score]
            weighted_deficiency += weight * deficiency
            weighted_total += weight

        if not any_answered or weighted_total == 0:
            return None
        return weighted_deficiency / weighted_total

    def _recalculate(self):
        ratios = []
        total_items = 0
        answered_items = 0

        for s_idx in range(len(SECTIONS)):
            total_items += len(self.section_vars[s_idx])
            answered_items += len([v for v in self.section_vars[s_idx] if v.get() != -1])

            ratio = self._section_deficiency_ratio(s_idx)
            lbl = self.section_total_labels[s_idx]
            if ratio is not None:
                ratios.append(ratio)
                pct = ratio * 100
                bg, fg = deficiency_color(pct)
                lbl.config(text=f"{pct:.1f}%", bg=bg, fg=fg)
            else:
                lbl.config(text="—", bg="#E8E8E8", fg="#333333")

            alert_widgets = self.section_alert_labels[s_idx]
            if alert_widgets:
                alert_row, pct_value_lbl, alert_lbl = alert_widgets
                critical_idx = CRITICAL_ITEM_INDICES[s_idx]

                answered_crit_vals = [self.section_vars[s_idx][i].get() for i in critical_idx if
                                      self.section_vars[s_idx][i].get() != -1]

                if not answered_crit_vals:
                    alert_row.config(bg="#F5F5F5")
                    pct_value_lbl.config(text="", bg="#F5F5F5", fg="#555555")
                    alert_lbl.config(text="", bg="#F5F5F5")
                elif -2 in answered_crit_vals:
                    bg_color, fg_color = ALERT_BG, ALERT_TXT
                    msg = "⚠ ΠΡΟΕΙΔΟΠΟΙΗΣΗ: Κρίσιμο κριτήριο Μη Εφαρμόσιμο (μέγιστη ποινή)"
                    alert_row.config(bg=bg_color)
                    pct_value_lbl.config(text="", bg=bg_color, fg=fg_color)
                    alert_lbl.config(text=msg, fg=fg_color, bg=bg_color)
                elif 0 in answered_crit_vals:
                    bg_color, fg_color = ORANGE_BG, ORANGE_TXT
                    msg = "⚠ Προσοχή: Κρίσιμο κριτήριο σε Σοβαρή Μη Συμμόρφωση"
                    alert_row.config(bg=bg_color)
                    pct_value_lbl.config(text="", bg=bg_color, fg=fg_color)
                    alert_lbl.config(text=msg, fg=fg_color, bg=bg_color)
                elif 1 in answered_crit_vals:
                    bg_color, fg_color = WARN_YELLOW_BG, WARN_YELLOW_TXT
                    msg = "⚠ Προσοχή: Κρίσιμο κριτήριο σε Ήπια Μη Συμμόρφωση"
                    alert_row.config(bg=bg_color)
                    pct_value_lbl.config(text="", bg=bg_color, fg=fg_color)
                    alert_lbl.config(text=msg, fg=fg_color, bg=bg_color)
                else:
                    alert_row.config(bg="#F5F5F5")
                    pct_value_lbl.config(text="", bg="#F5F5F5", fg="#555555")
                    alert_lbl.config(text="", bg="#F5F5F5")

        self.progress_lbl.config(text=f"Απαντημένα: {answered_items}/{total_items}")
        progress_pct = (answered_items / total_items) * 100 if total_items > 0 else 0
        self.progress_bar['value'] = progress_pct

        if answered_items == total_items:
            self.progress_lbl.config(fg=GOOD_GREEN_TXT, font=("Segoe UI", 10, "bold"))
        else:
            self.progress_lbl.config(fg="black", font=("Segoe UI", 10))

        # --- Υπολογισμός Τελικού Σκορ και Ποινής ---
        if ratios:
            base_x_norm = (sum(ratios) / len(ratios)) * 100
        else:
            base_x_norm = 0.0

        penalty_applied = 0.0

        if self.install_type_var.get() == "Κτηνοτροφική Εγκατάσταση" and self.history_loaded.get():
            prev_score_pct = self.history_prev_score.get() * 100
            penalty_pct, applies, note = self._compute_history_penalty()
            penalty_applied = penalty_pct

            if applies:
                self.hist_penalty_frame.config(bg="#E6F2FF")
                self.hist_penalty_lbl.config(
                    text=f"ΕΝΕΡΓΟΠΟΙΗΜΕΝΟ: Προηγούμενη Μη Συμμόρφωση: {prev_score_pct:.2f}% | "
                         f"{note} | Προστιθέμενη Ποινή: +{penalty_applied:.2f}%",
                    bg="#E6F2FF", fg="#1F497D", font=("Segoe UI", 10, "bold")
                )
            else:
                self.hist_penalty_frame.config(bg="#F0F0F0")
                self.hist_penalty_lbl.config(
                    text=f"Προηγούμενη Μη Συμμόρφωση: {prev_score_pct:.2f}% | {note}",
                    bg="#F0F0F0", fg="#666666", font=("Segoe UI", 10, "italic")
                )
        else:
            penalty_applied = 0.0
            self.hist_penalty_frame.config(bg="#DDDDDD")
            if self.install_type_var.get() == "Πρόχειρο Κατάλυμα":
                self.hist_penalty_lbl.config(
                    text="Προηγούμενη Συμμόρφωση: Απενεργοποιημένο (Πρόχειρο Κατάλυμα - Ποινή: 0.00%)", bg="#DDDDDD",
                    fg="#666666", font=("Segoe UI", 10, "italic"))
            else:
                self.hist_penalty_lbl.config(
                    text="Προηγούμενη Συμμόρφωση: Δεν έχει φορτωθεί/εισαχθεί ιστορικό (Ποινή: 0.00%)", bg="#DDDDDD",
                    fg="#666666", font=("Segoe UI", 10, "italic"))

        final_x_norm = min(100.0, base_x_norm + penalty_applied)

        if ratios:
            bg, fg = deficiency_color(final_x_norm)
            self.pct_label.config(text=f"{final_x_norm:.1f}%", bg=bg, fg=fg)
        else:
            self.pct_label.config(text="0.0%", bg=GOOD_GREEN_BG, fg=GOOD_GREEN_TXT)

    #-------------------------------------- ΕΠΙΚΥΡΩΣΗ & ΓΕΝΙΚΕΣ ΕΝΕΡΓΕΙΕΣ ΦΟΡΜΑΣ --------------------------------------#
    def validate_data(self):
        farm = self.farm_name_var.get().strip()
        inspector = self.inspector_var.get().strip()
        date_str = self.date_var.get().strip()

        if not farm:
            messagebox.showwarning("Ελλιπή Στοιχεία", "Παρακαλώ συμπληρώστε την Εκμετάλλευση / Κωδικό.")
            return False
        if not inspector:
            messagebox.showwarning("Ελλιπή Στοιχεία", "Παρακαλώ συμπληρώστε το όνομα του Ελεγκτή/τριας.")
            return False

        if not re.match(r"^\d{2}/\d{2}/\d{4}$", date_str):
            messagebox.showwarning("Λάθος Μορφή",
                                   "Η ημερομηνία πρέπει να είναι στη μορφή ΗΗ/ΜΜ/ΕΕΕΕ (π.χ. 25/06/2026).")
            return False

        for s_idx, vars_for_section in enumerate(self.section_vars):
            for v_idx, var in enumerate(vars_for_section):
                if var.get() == -1:
                    messagebox.showwarning("Αναπάντητα Κριτήρια",
                                           f"Δεν έχουν βαθμολογηθεί όλα τα κριτήρια.\nΕκκρεμότητα: Ενότητα {s_idx + 1}, Κριτήριο {v_idx + 1}.")
                    return False
        return True

    def sanitize_filename(self, name):
        return re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_")

    def on_closing(self):
        if self.unsaved_changes:
            if not messagebox.askyesno("Έξοδος",
                                       "Έχετε μη αποθηκευμένες αλλαγές. Είστε σίγουροι ότι θέλετε να κλείσετε την εφαρμογή;"):
                return
        self.root.destroy()

    def clear_form(self):
        if self.unsaved_changes:
            if not messagebox.askyesno("Επιβεβαίωση", "Έχετε μη αποθηκευμένες αλλαγές. Να γίνει οριστικός καθαρισμός;"):
                return
        self.farm_name_var.set("")
        self.inspector_var.set("")
        self.date_var.set(datetime.now().strftime("%d/%m/%Y"))

        for section_vars in self.section_vars:
            for v in section_vars:
                v.set(-1)

        self.unsaved_changes = False
        self._reset_historical_data()

    # --- ΕΞΑΓΩΓΗ EXCEL ---
    #----------------------------------------------------- EXCEL ------------------------------------------------------#
    def export_to_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Λείπει βιβλιοθήκη", "Η βιβλιοθήκη 'openpyxl' δεν είναι εγκατεστημένη.")
            return

        if not self.validate_data():
            return

        farm_safe = self.sanitize_filename(self.farm_name_var.get())
        formatted_date = datetime.now().strftime('%d-%m-%Y')
        default_name = f"Checklist_Biosecurity_{farm_safe}_{formatted_date}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Αποθήκευση checklist ως Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return

        try:
            self._write_excel(path)
            self.unsaved_changes = False
            messagebox.showinfo("Επιτυχία", f"Το checklist αποθηκεύτηκε επιτυχώς:\n{path}\n\n")
        except Exception as exc:
            messagebox.showerror("Σφάλμα εξαγωγής", f"Δεν ήταν δυνατή η αποθήκευση του αρχείου:\n{exc}")

    def _write_excel(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Λίστα Ελέγχου"

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        widths = {"A": 6, "B": 70, "C": 16, "D": 18, "E": 20, "F": 18, "G": 20}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        ws["A1"] = "ΥΠΟΥΡΓΕΙΟ ΑΓΡΟΤΙΚΗΣ ΑΝΑΠΤΥΞΗΣ ΚΑΙ ΤΡΟΦΙΜΩΝ - ΓΕΝΙΚΗ ΔΙΕΥΘΥΝΣΗ ΚΤΗΝΙΑΤΡΙΚΗΣ"
        ws["A1"].font = Font(bold=True, size=14, color="1F497D")
        ws.merge_cells("A1:G1")

        ws["B2"] = f"ΛΙΣΤΑ ΕΛΕΓΧΟΥ ΒΙΟΠΡΟΦΥΛΑΞΗΣ: {self.install_type_var.get().upper()}"
        ws["B2"].font = Font(bold=True, size=12)

        instructions_text = (
            "Οδηγίες: Συμμόρφωση, Ήπια Μη Συμμόρφωση, Σοβαρή Μη Συμμόρφωση, Μη Εφαρμόσιμο. Το ποσοστό στη στήλη "
            "\"Βαθμολογία\" δείχνει πόσο συνεισφέρει το κάθε κριτήριο στο Σύνολο Αστοχίας της ενότητας.\n"
            "Τα κριτήρια με 🔺 και έντονη γραφή (bold) έχουν αυξημένη βαρύτητα. Το Μη Εφαρμόσιμο μετράει ως η "
            "χειρότερη δυνατή βαθμολογία."
        )

        ws.merge_cells("B3:G3")
        ws["B3"] = instructions_text
        ws["B3"].font = Font(size=9, italic=True)
        ws["B3"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[3].height = 58

        ws["B4"] = (
            f"Εκμετάλλευση: {self.farm_name_var.get()}    "
            f"Ελεγκτής/τρια: {self.inspector_var.get()}    "
            f"Ημερομηνία: {self.date_var.get()}"
        )
        ws["B4"].font = Font(size=12, bold=True, color="555555")

        headers = ["Α/Α", "Κριτήριο Ελέγχου", "ΣΥΜΜΟΡΦΩΣΗ", "ΗΠΙΑ ΜΗ ΣΥΜΜΟΡΦΩΣΗ",
                   "ΣΟΒΑΡΗ ΜΗ ΣΥΜΜΟΡΦΩΣΗ", "ΜΗ ΕΦΑΡΜΟΣΙΜΟ", "Βαθμολογία (%)"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        row_num = 6
        an_a = 1
        section_total_rows = []
        flat_data_rows = []
        ratios = []

        for s_idx, section in enumerate(SECTIONS):
            ws.merge_cells(f"A{row_num}:G{row_num}")
            cell = ws.cell(row=row_num, column=1, value=section["title"])
            cell.font = Font(bold=True, size=11, color="1F497D")
            cell.fill = PatternFill("solid", fgColor="DCE6F1")
            cell.alignment = Alignment(horizontal="left")
            row_num += 1

            first_item_row = row_num
            scores = [v.get() for v in self.section_vars[s_idx]]
            critical_idx = CRITICAL_ITEM_INDICES[s_idx]

            total_weight_section = sum(
                (CRITICAL_WEIGHT if i in critical_idx else NORMAL_WEIGHT)
                for i in range(len(section["items"]))) or 1.0

            for item_idx, (item_text, score) in enumerate(zip(section["items"], scores)):
                is_critical = item_idx in critical_idx
                display_text = ("🔺 " + item_text) if is_critical else item_text

                ws.cell(row=row_num, column=1, value=an_a).border = border
                b_cell = ws.cell(row=row_num, column=2, value=display_text)
                b_cell.border = border
                b_cell.alignment = Alignment(wrap_text=True, vertical="top")
                if is_critical:
                    b_cell.font = Font(bold=True)
                for col in (3, 4, 5, 6):
                    c = ws.cell(row=row_num, column=col)
                    c.border = border
                    c.alignment = Alignment(horizontal="center")

                weight = CRITICAL_WEIGHT if is_critical else NORMAL_WEIGHT
                raw_deficiency = DEFICIENCY_BY_SCORE[score]

                tick_col = {2: 3, 1: 4, 0: 5, -2: 6}[score]

                if is_critical and score == 1:
                    tick_bg, tick_fg = WARN_YELLOW_BG.lstrip("#"), WARN_YELLOW_TXT.lstrip("#")
                elif is_critical and score == 0:
                    tick_bg, tick_fg = ORANGE_BG.lstrip("#"), ORANGE_TXT.lstrip("#")
                elif score == -2:
                    tick_bg, tick_fg = BAD_RED_BG.lstrip("#"), BAD_RED_TXT.lstrip("#")
                else:
                    tick_bg, tick_fg = None, "000000"

                tick_cell = ws.cell(row=row_num, column=tick_col, value="✓")
                tick_cell.font = Font(bold=True, size=12, color=tick_fg)
                if tick_bg:
                    tick_cell.fill = PatternFill("solid", fgColor=tick_bg)
                tick_cell.alignment = Alignment(horizontal="center")
                tick_cell.border = border

                row_contribution = (weight * raw_deficiency) / total_weight_section

                f_cell = ws.cell(row=row_num, column=7, value=row_contribution)
                f_cell.number_format = "0.0%"
                f_cell.border = border
                f_cell.alignment = Alignment(horizontal="center")
                if is_critical:
                    f_cell.font = Font(bold=True)

                flat_data_rows.append([
                    self.date_var.get(), self.farm_name_var.get(), self.inspector_var.get(),
                    section["title"], f"C{an_a:02d}", item_text,
                    "Μη Εφαρμόσιμο" if score == -2 else score, raw_deficiency, "Ναι" if is_critical else "Όχι"
                ])
                an_a += 1
                row_num += 1

            last_item_row = row_num - 1

            ws.cell(row=row_num, column=2, value="Σύνολο Συμμόρφωσης Ενότητας %").alignment = Alignment(
                horizontal="right")
            formula = f"=SUM(G{first_item_row}:G{last_item_row})"
            total_cell = ws.cell(row=row_num, column=7, value=formula)
            total_cell.number_format = "0.0%"

            for col in range(1, 8):
                c = ws.cell(row=row_num, column=col)
                c.fill = PatternFill("solid", fgColor="E8E8E8")
                c.font = Font(bold=True, color="333333")
                c.border = border

            ratio = self._section_deficiency_ratio(s_idx)
            if ratio is not None:
                ratios.append(ratio)
            sect_bg, sect_fg = deficiency_color((ratio * 100) if ratio is not None else 0.0)
            total_cell.fill = PatternFill("solid", fgColor=sect_bg.lstrip("#"))
            total_cell.font = Font(bold=True, color=sect_fg.lstrip("#"))
            total_cell.alignment = Alignment(horizontal="center")
            section_total_rows.append(row_num)
            row_num += 1

            critical_vals_in_section = [scores[i] for i in critical_idx]
            if -2 in critical_vals_in_section:
                warn_bg, warn_fg = ALERT_BG, ALERT_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο Μη Εφαρμόσιμο (μέγιστη ποινή)"
            elif 0 in critical_vals_in_section:
                warn_bg, warn_fg = ORANGE_BG, ORANGE_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο σε Σοβαρή Μη Συμμόρφωση"
            elif 1 in critical_vals_in_section:
                warn_bg, warn_fg = WARN_YELLOW_BG, WARN_YELLOW_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο σε Ήπια Μη Συμμόρφωση"
            else:
                warn_bg = None

            if warn_bg:
                ws.merge_cells(f"A{row_num}:G{row_num}")
                wcell = ws.cell(row=row_num, column=1, value=warn_msg)
                wcell.font = Font(bold=True, size=10, color=warn_fg.lstrip("#"))
                wcell.fill = PatternFill("solid", fgColor=warn_bg.lstrip("#"))
                wcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws.row_dimensions[row_num].height = 18
                row_num += 1

            row_num += 1

        grand_total_row = row_num
        f_refs = "+".join(f"G{r}" for r in section_total_rows)
        base_formula = f"({f_refs})/{len(section_total_rows)}"
        base_x_norm_frac = (sum(ratios) / len(ratios)) if ratios else 0.0

        penalty = 0.0
        if self.history_loaded.get() and self.install_type_var.get() == "Κτηνοτροφική Εγκατάσταση":
            ws.cell(row=grand_total_row, column=2, value="Αρχική Μη Συμμόρφωση (Χωρίς Ποινή):").font = Font(bold=True,
                                                                                                            size=10)
            base_cell = ws.cell(row=grand_total_row, column=7, value=f"={base_formula}")
            base_cell.number_format = "0.00%"
            base_bg, base_fg = deficiency_color(base_x_norm_frac * 100)
            base_cell.fill = PatternFill("solid", fgColor=base_bg.lstrip("#"))
            base_cell.font = Font(bold=True, color=base_fg.lstrip("#"))

            penalty_row = grand_total_row + 1
            prev_score = self.history_prev_score.get()
            penalty_pct, applies, note = self._compute_history_penalty()
            penalty = penalty_pct / 100.0

            if applies:
                ws.cell(row=penalty_row, column=2,
                        value=f"ΠΡΟΣΤΙΘΕΜΕΝΗ ΠΟΙΝΗ ΙΣΤΟΡΙΚΟΥ (Προηγ. έλεγχος: {prev_score:.2%}, {note}):").font = Font(
                    bold=True, size=10, color="9C0006")
                p_cell = ws.cell(row=penalty_row, column=7, value=penalty)
                p_cell.number_format = "0.00%"
                p_cell.font = Font(bold=True, color="9C0006")
                grand_total_row += 2
            else:
                grand_total_row += 1

        final_frac = min(1.0, base_x_norm_frac + penalty)
        final_bg, final_fg = deficiency_color(final_frac * 100)

        ws.cell(row=grand_total_row, column=2,
               value="ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ ΒΑΡΥΤΗΤΑΣ ΜΗ ΣΥΜΜΟΡΦΩΣΗΣ [0-1] norm%").font = Font(bold=True, size=12)
        norm_cell = ws.cell(row=grand_total_row, column=7, value=f"=MIN(1.0, {base_formula} + {penalty})")
        norm_cell.font = Font(bold=True, size=13, color=final_fg.lstrip("#"))
        norm_cell.fill = PatternFill("solid", fgColor=final_bg.lstrip("#"))
        norm_cell.number_format = "0.00%"

        legend_row = grand_total_row + 2
        ws.row_dimensions[legend_row].height = 30
        legend_label_cell = ws.cell(row=legend_row, column=2, value="Υπόμνημα:")
        legend_label_cell.font = Font(size=8, bold=True, color="555555")
        legend_label_cell.alignment = Alignment(horizontal="right", vertical="center")

        legend_items = [
            (3, "0-30%\nΣυμμόρφωση", GOOD_GREEN_BG, GOOD_GREEN_TXT),
            (4, "30-50%\nΉπια Μη Συμμόρφωση", WARN_YELLOW_BG, WARN_YELLOW_TXT),
            (5, "50-90%\nΣοβαρή Μη Συμμόρφωση", ORANGE_BG, ORANGE_TXT),
            (6, "90-100%\nΜη Εφαρμόσιμο", BAD_RED_BG, BAD_RED_TXT),
        ]
        for col, text, bg_hex, fg_hex in legend_items:
            c = ws.cell(row=legend_row, column=col, value=text)
            c.font = Font(size=8, bold=True, color=fg_hex.lstrip("#"))
            c.fill = PatternFill("solid", fgColor=bg_hex.lstrip("#"))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

        credit_row = legend_row + 2
        ws.cell(row=credit_row, column=1, value="Created by © 2026 Aggelos A. Kaptanis").font = Font(size=7,
                                                                                                     italic=True,
                                                                                                     color="AAAAAA")

        ws_meta = wb.create_sheet("Meta")
        ws_meta.sheet_state = 'hidden'
        ws_meta["A1"] = "Κωδικός Εκμετάλλευσης"
        ws_meta["B1"] = self.farm_name_var.get()
        ws_meta["A2"] = "Ημερομηνία"
        ws_meta["B2"] = self.date_var.get()
        ws_meta["A3"] = "Τελικό Ποσοστό Μη Συμμόρφωσης"
        ws_meta["B3"] = final_frac

        ws2 = wb.create_sheet("Δεδομένα_BI")
        ws2.sheet_state = 'hidden'

        bi_headers = ["Ημερομηνία", "Κωδ. Εκμετάλλευσης", "Ελεγκτής", "Ενότητα", "ID Κριτηρίου", "Περιγραφή Κριτηρίου",
                      "Βαθμολογία (2,1,0)", "Δείκτης Αστοχίας (0, 0.5, 1)", "Κρίσιμο Κριτήριο"]
        ws2.append(bi_headers)
        for col_idx in range(1, len(bi_headers) + 1):
            c = ws2.cell(row=1, column=col_idx)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DCE6F1")
        for r_data in flat_data_rows:
            ws2.append(r_data)

        for col_letter, width in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I"],
                                     [12, 20, 20, 40, 15, 60, 20, 25, 15]):
            ws2.column_dimensions[col_letter].width = width

        wb.save(path)

    # --- ΕΞΑΓΩΓΗ PDF ---
    #------------------------------------------------------ PDF -------------------------------------------------------#
    def export_to_pdf(self):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Λείπει βιβλιοθήκη", "Η βιβλιοθήκη 'reportlab' δεν είναι εγκατεστημένη.")
            return

        if not self.validate_data():
            return

        farm_safe = self.sanitize_filename(self.farm_name_var.get())
        formatted_date = datetime.now().strftime('%d-%m-%Y')
        default_name = f"Checklist_Biosecurity_{farm_safe}_{formatted_date}.pdf"

        path = filedialog.asksaveasfilename(
            title="Αποθήκευση checklist ως PDF",
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF files", "*.pdf")],
        )
        if not path:
            return

        try:
            self._write_pdf(path)
            self.unsaved_changes = False
            messagebox.showinfo("Επιτυχία", f"Το checklist αποθηκεύτηκε επιτυχώς:\n{path}\n\n")
        except Exception as exc:
            messagebox.showerror("Σφάλμα εξαγωγής", f"Δεν ήταν δυνατή η αποθήκευση του αρχείου:\n{exc}")

    def _write_pdf(self, path):
        _register_pdf_fonts()

        style_normal = ParagraphStyle("normal_el", fontName=PDF_FONT_NAME, fontSize=8.5, leading=11, alignment=TA_LEFT)
        style_normal_bold = ParagraphStyle("normal_el_bold", fontName=PDF_FONT_NAME_BOLD, fontSize=8.5, leading=11,
                                           alignment=TA_LEFT)
        style_title = ParagraphStyle("title_el", fontName=PDF_FONT_NAME_BOLD, fontSize=13, leading=16,
                                     alignment=TA_CENTER, textColor=rl_colors.HexColor(SECTION_BLUE_TXT))
        style_subtitle = ParagraphStyle("subtitle_el", fontName=PDF_FONT_NAME_BOLD, fontSize=11, leading=14,
                                        alignment=TA_CENTER)
        style_info = ParagraphStyle("info_el", fontName=PDF_FONT_NAME_BOLD, fontSize=10, leading=14,
                                    textColor=rl_colors.HexColor("#555555"))
        style_instr = ParagraphStyle("instr_el", fontName=PDF_FONT_NAME, fontSize=7.5, leading=10,
                                     textColor=rl_colors.HexColor("#555555"))
        style_legend_caption = ParagraphStyle("legend_caption_el", fontName=PDF_FONT_NAME_BOLD, fontSize=7.5,
                                              leading=10, textColor=rl_colors.HexColor("#555555"))
        style_section_title = ParagraphStyle("section_el", fontName=PDF_FONT_NAME_BOLD, fontSize=10, leading=13,
                                             textColor=rl_colors.HexColor(SECTION_BLUE_TXT))
        style_credit = ParagraphStyle("credit_el", fontName=PDF_FONT_NAME, fontSize=6.5,
                                      textColor=rl_colors.HexColor("#AAAAAA"))

        doc = SimpleDocTemplate(
            path, pagesize=landscape(A4),
            leftMargin=14 * mm, rightMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
            title="Λίστα Ελέγχου Βιοπροφύλαξης",
        )

        elements = []
        elements.append(
            Paragraph("ΥΠΟΥΡΓΕΙΟ ΑΓΡΟΤΙΚΗΣ ΑΝΑΠΤΥΞΗΣ ΚΑΙ ΤΡΟΦΙΜΩΝ - ΓΕΝΙΚΗ ΔΙΕΥΘΥΝΣΗ ΚΤΗΝΙΑΤΡΙΚΗΣ", style_title))
        elements.append(Paragraph(f"ΛΙΣΤΑ ΕΛΕΓΧΟΥ ΒΙΟΠΡΟΦΥΛΑΞΗΣ: {self.install_type_var.get().upper()}", style_subtitle))
        elements.append(Spacer(1, 6))

        elements.append(Paragraph(
            "Οδηγίες: Το ποσοστό δείχνει πόσο συνεισφέρει το κριτήριο στο Σύνολο Αστοχίας. Τα κριτήρια με * έχουν "
            "αυξημένη βαρύτητα. Το Μη Εφαρμόσιμο μετράει ως η χειρότερη δυνατή βαθμολογία.",
            style_instr))
        elements.append(Spacer(1, 6))

        info_text = (f"Εκμετάλλευση: {self.farm_name_var.get()} &nbsp;&nbsp;&nbsp;&nbsp; "
                     f"Ελεγκτής/τρια: {self.inspector_var.get()} &nbsp;&nbsp;&nbsp;&nbsp; "
                     f"Ημερομηνία: {self.date_var.get()}")
        elements.append(Paragraph(info_text, style_info))
        elements.append(Spacer(1, 10))

        col_widths = [11 * mm, 118 * mm, 16 * mm, 20 * mm, 22 * mm, 18 * mm, 20 * mm]
        header_row_cells = ["Α/Α", "Κριτήριο Ελέγχου", "ΣΥΜΜ.", "ΗΠΙΑ ΜΗ\nΣΥΜΜ.", "ΣΟΒΑΡΗ ΜΗ\nΣΥΜΜ.",
                            "ΜΗ ΕΦ.", "Βαθμ.\n%"]
        ratios = []
        an_a = 1

        for s_idx, section in enumerate(SECTIONS):
            elements.append(Paragraph(section["title"], style_section_title))
            elements.append(Spacer(1, 3))

            critical_idx = CRITICAL_ITEM_INDICES[s_idx]
            total_weight_section = sum(
                (CRITICAL_WEIGHT if i in critical_idx else NORMAL_WEIGHT)
                for i in range(len(section["items"]))) or 1.0
            table_data = [header_row_cells]
            extra_style_cmds = []

            for item_idx, item_text in enumerate(section["items"]):
                score = self.section_vars[s_idx][item_idx].get()
                is_critical = item_idx in critical_idx
                is_na = (score == -2)
                marker = ("* ") if is_critical else ""
                label_text = marker + item_text
                text_style = style_normal_bold if is_critical else style_normal
                criterion_p = Paragraph(label_text, text_style)

                weight = CRITICAL_WEIGHT if is_critical else NORMAL_WEIGHT
                raw_deficiency = DEFICIENCY_BY_SCORE[score]
                row_contribution_pct = (weight * raw_deficiency) / total_weight_section * 100

                tick = "X"
                c_ok = tick if score == 2 else ""
                c_mid = tick if score == 1 else ""
                c_bad = tick if score == 0 else ""
                c_na = tick if is_na else ""

                table_data.append(
                    [str(an_a), criterion_p, c_ok, c_mid, c_bad, c_na, f"{row_contribution_pct:.1f}%"])
                row_i = len(table_data) - 1

                if is_critical and score == 1:
                    extra_style_cmds.append(("BACKGROUND", (3, row_i), (3, row_i), rl_colors.HexColor(WARN_YELLOW_BG)))
                    extra_style_cmds.append(("TEXTCOLOR", (3, row_i), (3, row_i), rl_colors.HexColor(WARN_YELLOW_TXT)))
                elif is_critical and score == 0:
                    extra_style_cmds.append(("BACKGROUND", (4, row_i), (4, row_i), rl_colors.HexColor(ORANGE_BG)))
                    extra_style_cmds.append(("TEXTCOLOR", (4, row_i), (4, row_i), rl_colors.HexColor(ORANGE_TXT)))
                elif is_na:
                    extra_style_cmds.append(("BACKGROUND", (5, row_i), (5, row_i), rl_colors.HexColor(BAD_RED_BG)))
                    extra_style_cmds.append(("TEXTCOLOR", (5, row_i), (5, row_i), rl_colors.HexColor(BAD_RED_TXT)))

                an_a += 1

            ratio = self._section_deficiency_ratio(s_idx)
            pct = ratio * 100 if ratio is not None else 0.0
            if ratio is not None:
                ratios.append(ratio)

            table_data.append(["", "Σύνολο Συμμόρφωσης Ενότητας %", "", "", "", "", f"{pct:.1f}%"])
            total_row_i = len(table_data) - 1

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            sect_bg, sect_fg = deficiency_color(pct)
            style_cmds = [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_NAME_BOLD), ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(HEADER_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 1), (0, -1), "CENTER"), ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#B0B0B0")),
                ("BACKGROUND", (0, total_row_i), (-1, total_row_i), rl_colors.HexColor("#E8E8E8")),
                ("FONTNAME", (0, total_row_i), (-1, total_row_i), PDF_FONT_NAME_BOLD),
                ("SPAN", (1, total_row_i), (5, total_row_i)), ("ALIGN", (1, total_row_i), (1, total_row_i), "RIGHT"),
                ("BACKGROUND", (6, total_row_i), (6, total_row_i), rl_colors.HexColor(sect_bg)),
                ("TEXTCOLOR", (6, total_row_i), (6, total_row_i), rl_colors.HexColor(sect_fg)),
            ]
            style_cmds.extend(extra_style_cmds)
            table.setStyle(TableStyle(style_cmds))
            elements.append(table)

            critical_scores_in_section = [self.section_vars[s_idx][i].get() for i in critical_idx]
            if -2 in critical_scores_in_section:
                warn_bg, warn_fg = ALERT_BG, ALERT_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο Μη Εφαρμόσιμο (μέγιστη ποινή)"
            elif 0 in critical_scores_in_section:
                warn_bg, warn_fg = ORANGE_BG, ORANGE_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο σε Σοβαρή Μη Συμμόρφωση"
            elif 1 in critical_scores_in_section:
                warn_bg, warn_fg = WARN_YELLOW_BG, WARN_YELLOW_TXT
                warn_msg = "⚠ Δείκτης Συναγερμού Κρίσιμων: Κρίσιμο κριτήριο σε Ήπια Μη Συμμόρφωση"
            else:
                warn_bg = None

            if warn_bg:
                warn_style = ParagraphStyle(
                    f"warn_el_{s_idx}", fontName=PDF_FONT_NAME_BOLD, fontSize=9,
                    textColor=rl_colors.HexColor(warn_fg), backColor=rl_colors.HexColor(warn_bg),
                    alignment=TA_LEFT, borderPadding=5,
                )
                elements.append(Spacer(1, 3))
                elements.append(Paragraph(warn_msg, warn_style))

            elements.append(Spacer(1, 10))

        base_x_norm = (sum(ratios) / len(ratios)) * 100 if ratios else 0.0
        penalty_applied = 0.0

        if self.history_loaded.get() and self.install_type_var.get() == "Κτηνοτροφική Εγκατάσταση":
            prev_score_pct = self.history_prev_score.get() * 100
            penalty_pct, applies, note = self._compute_history_penalty()
            penalty_applied = penalty_pct

            hist_style = ParagraphStyle("hist_el", fontName=PDF_FONT_NAME_BOLD, fontSize=11,
                                        textColor=rl_colors.HexColor("#9C0006"), alignment=TA_CENTER)
            base_bg, base_fg = deficiency_color(base_x_norm)
            elements.append(Paragraph(f"ΑΡΧΙΚΗ ΜΗ ΣΥΜΜΟΡΦΩΣΗ: {base_x_norm:.2f}%",
                                      ParagraphStyle("s1", parent=hist_style, textColor=rl_colors.HexColor(base_fg),
                                                     backColor=rl_colors.HexColor(base_bg), borderPadding=4)))
            elements.append(Spacer(1, 4))
            if applies:
                elements.append(Paragraph(
                    f"+ ΠΟΙΝΗ ΙΣΤΟΡΙΚΟΥ (Προηγούμενη {prev_score_pct:.2f}%, {note}): +{penalty_applied:.2f}%",
                    hist_style))
                elements.append(Spacer(1, 8))

        final_x_norm = min(100.0, base_x_norm + penalty_applied)
        bg, fg = deficiency_color(final_x_norm)

        final_style = ParagraphStyle(
            "final_el", fontName=PDF_FONT_NAME_BOLD, fontSize=13,
            textColor=rl_colors.HexColor(fg), backColor=rl_colors.HexColor(bg),
            alignment=TA_CENTER, borderPadding=8,
        )
        elements.append(Paragraph(f"ΤΕΛΙΚΟΣ ΔΕΙΚΤΗΣ ΒΑΡΥΤΗΤΑΣ ΜΗ ΣΥΜΜΟΡΦΩΣΗΣ [0-1] norm%: {final_x_norm:.1f}%",
                                  final_style))

        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Κλίμακα ερμηνείας ποσοστού (%) αστοχίας:", style_legend_caption))
        elements.append(Spacer(1, 2))
        elements.append(_build_legend_drawing(doc.width * 0.42))

        elements.append(Spacer(1, 14))
        elements.append(Paragraph("Created by © 2026 Aggelos A. Kaptanis", style_credit))

        doc.build(elements)


#--------------------------------------------- ΕΚΚΙΝΗΣΗ ΕΦΑΡΜΟΓΗΣ (MAIN) ----------------------------------------------#
def main():
    root = tk.Tk()

    try:
        root.iconbitmap(resource_path(ICON_FILENAME))
    except Exception:
        pass

    try:
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("Green.Horizontal.TProgressbar", background="#4CAF50", troughcolor="#E8E8E8")
    except Exception:
        pass

    app = BiosafetyChecklistApp(root)
    root.mainloop()


if __name__ == "__main__":
    show_splash_and_start(main)